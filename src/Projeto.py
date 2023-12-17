import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

def esperar_elemento(driver, by, valor, timeout=10):
    try:
        # Aguarda até que o elemento esteja presente na página
        elemento = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, valor))
        )
        return elemento
    except TimeoutException:
        print(f"Tempo de espera excedido para o elemento {by}: {valor}")
        return None

def extrair_dados(driver, xpaths):
    dados = []
    for xpath in xpaths:
        # Extrai dados dos elementos correspondentes aos XPaths fornecidos
        elemento = esperar_elemento(driver, By.XPATH, xpath)
        dados.append(elemento.text if elemento else None)
    return dados

# Configurando o WebDriver
servico = Service(ChromeDriverManager().install())
arquivoExcel = 'Demo.xlsx'
urlDetran = "https://consultas.detrannet.sc.gov.br/servicos/consultaveiculo.asp"

# Atribui ao data frame o arquivo excel
df = pd.read_excel(arquivoExcel)

# Inicialização do WebDriver, Workbook e Worksheet
chrome = webdriver.Chrome(service=servico)
wb = load_workbook(arquivoExcel)
ws = wb.active

# Percorre as linhas do DataFrame
for index, row in df.iterrows():
    # Aqui é verificado se a linha contém informações
    if pd.isna(row["PLACA"]):
        print("Não há mais informações. Interrompendo o loop.")
        break

    chrome.get(urlDetran)

    elementoPlaca = esperar_elemento(chrome, By.XPATH, '//*[@id="placa"]')
    elementoRenavam = esperar_elemento(chrome, By.XPATH, '//*[@id="renavam"]')

    # Preenche os campos com os valores do DataFrame
    elementoPlaca.send_keys(row["PLACA"])
    elementoRenavam.send_keys(row["RENAVAM"])

    # Clica no botão de consulta
    chrome.find_element(By.XPATH, '//*[@id="form1"]/table[2]/tbody/tr[4]/td/fieldset/table/tbody/tr[4]/td/button').click()

    try:
        # Aguarda até que o elemento da tabela esteja presente na página
        elementoTabela = esperar_elemento(chrome, By.XPATH, '//*[@id="div_servicos_09"]')
        if not elementoTabela:
            print("ElementoTabela não encontrado. Verifique o XPath ou aguarde a página carregar completamente.")
            break
    except NoSuchElementException:
        print("ElementoTabela não encontrado. Verifique o XPath ou aguarde a página carregar completamente.")
        break

    xpaths = [
        '//*[@id="div_servicos_09"]/table/tbody/tr[2]/td[4]/table/tbody/tr[1]/td',
        '//*[@id="div_servicos_09"]/table/tbody/tr[3]/td[4]/table/tbody/tr[1]/td',
        '//*[@id="div_servicos_09"]/table/tbody/tr[4]/td[4]/table/tbody/tr[1]/td',
        '//*[@id="div_servicos_09"]/table/tbody/tr[5]/td[4]/table/tbody/tr[1]/td'
    ]

    # Encontra a próxima coluna vazia
    col_num = 1
    while ws.cell(row=index + 2, column=col_num).value is not None:
        col_num += 1

    # Extrai os dados e os adiciona à próxima coluna vazia na linha atual
    dados = extrair_dados(chrome, xpaths)
    for i, data in enumerate(dados):
        ws.cell(row=index + 2, column=col_num + i, value=data)

    # Salva os dados
    wb.save(arquivoExcel)

chrome.quit()
